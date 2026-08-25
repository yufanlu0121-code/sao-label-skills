"""Build a blind, stratified hand-annotation workbook for a validation sample.

TEMPLATE: adapt CODES, the column layouts, and the context-window rule to your schema.
The design machinery -- stratifying on the model's own labels, allocation with a floor
and a pooled rare stratum, inclusion probabilities computed at the finest allocation
level, and the blindness guarantees -- is domain-independent.

The extraction's labels are measurements with unknown error. This builds the instrument
for measuring that error: two stratified draws exported as an xlsx the coder types into,
with the model's labels withheld so the coding is not anchored.

Read `references/validation.md` before changing the sampling, and note two properties
that are easy to destroy:

- Stratifying on the model's own labels is efficient for estimating error rates
  conditional on predicted class, and valid ONLY if the inclusion probabilities reach
  the analysis. An unweighted analysis of this sample is wrong and nothing downstream
  will flag it.
- Anything whose *shape* depends on the model's output can leak it. The context window
  here is built without consulting the model's extracted sentence for exactly this
  reason -- see `references/pitfalls.md` #15.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Reused rather than reimplemented: locating a quoted sentence in a PDF-derived
# narrative fails on injected page headers and wrapped hyphens, which is documented at
# length in handcheck.py. A first version of this script searched the raw narrative and
# fell back for 17 of 150 documents that were in fact locatable.
from handcheck import normalize_punct, strip_furniture

BASE_DIR = Path(__file__).resolve().parent
NARRATIVES_PATH = BASE_DIR / "narratives.parquet"
DEFAULT_OUT_ROOT = Path(
    os.environ.get("SAO_OUT_ROOT", Path.home() / "sao_extract_output")
)
TEXT_COLUMNS = ("text", "section_relevant_comments")
RANDOM_STATE = 666

STRATUM_FIELDS = ["orientation", "object", "moment"]
RARE_THRESHOLD = 0.005
CELL_FLOOR = 40
CELL_CAP = 120
DOC_TARGET = 150

CODES = {
    "orientation": ["retrospective", "prospective", "both"],
    "object": ["exposure", "estimate_method", "data_quality", "external_party"],
    "moment": ["level", "uncertainty", "both"],
    "rmad_conclusion": ["yes", "no", "conditional", "not_stated"],
}

CONTEXT_CHARS = 200
RMAD_CONTEXT_CHARS = 600
# Used to centre the window when the model reported no RMAD sentence: the coder still
# has to decide whether the filing states a conclusion, and the head of the narrative
# is rarely where that is settled.
RMAD_KEYWORDS = (
    "material adverse deviation",
    "materially adverse",
    "risk of material adverse",
    "materiality standard",
)

STMT_COLUMNS = [
    "stmt_id",
    "filing_year",
    "prev_verbatim",
    "verbatim",
    "next_verbatim",
    "orientation",
    "object",
    "moment",
    "notes",
]
DOC_COLUMNS = [
    "filing_id",
    "filing_year",
    "rmad_sentence_context",
    "rmad_conclusion",
    "notes",
]
STMT_CODING = ["orientation", "object", "moment", "notes"]
DOC_CODING = ["rmad_conclusion", "notes"]

CODING_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
EXAMPLE_FILL = PatternFill("solid", fgColor="E2EFDA")

logging.basicConfig(
    level=logging.INFO, format="%(levelname)s %(message)s", stream=sys.stdout
)
logger = logging.getLogger(__name__)


def narrative_column(frame: pd.DataFrame) -> str:
    """Return whichever column holds the narrative text.

    Args:
        frame: The corpus frame.

    Returns:
        The column name.

    Raises:
        SystemExit: If no known text column is present.
    """
    for column in TEXT_COLUMNS:
        if column in frame.columns:
            return column
    raise SystemExit(f"No text column found; expected one of {TEXT_COLUMNS}.")


def xl_safe(value: Any) -> Any:
    """Strip characters Excel refuses to store.

    Narratives come from PDF text layers and carry control characters that openpyxl
    rejects outright with IllegalCharacterError. They are invisible in the source and
    meaningless in a spreadsheet cell, so they are removed rather than escaped.

    Args:
        value: Any cell value.

    Returns:
        The value, with illegal characters removed if it was a string.
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def squash(text: str) -> str:
    """Collapse whitespace to single spaces.

    Args:
        text: Any string.

    Returns:
        A new single-line string.
    """
    return " ".join(xl_safe(str(text)).split())


def clip(text: str, limit: int) -> str:
    """Truncate on a word boundary, appending an ellipsis when shortened.

    Args:
        text: The text to shorten.
        limit: Approximate character budget.

    Returns:
        A new string no longer than roughly `limit` characters.
    """
    flat = squash(text)
    if len(flat) <= limit:
        return flat
    cut = flat[:limit].rsplit(" ", 1)[0]
    return cut + " …"


def load_frames(out_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Explode the extractions into statement-level and document-level frames.

    `assemble.py` cannot supply these: it reads a per-model `raw/` directory and its
    `split_id()` parses `{naic}_{year}` from the filename stem, which raises on every
    current `{naic}_{year}_v1` key. The extractions are therefore read directly, as
    `report.py` does.

    Args:
        out_dir: Directory of extraction JSON files.

    Returns:
        A tuple of (statements, documents) frames.
    """
    corpus = pd.read_parquet(NARRATIVES_PATH).set_index("sao_id")
    text_col = narrative_column(corpus.reset_index())
    recovered = (
        corpus["_recovered"]
        if "_recovered" in corpus.columns
        else pd.Series(False, index=corpus.index)
    )

    statements: list[dict[str, Any]] = []
    documents: list[dict[str, Any]] = []
    for sao_id in corpus.index:
        record = json.loads((out_dir / f"{sao_id}.json").read_text(encoding="utf-8"))
        row = corpus.loc[sao_id]
        items = record.get("statements", [])
        verbatims = [squash(s.get("verbatim", "")) for s in items]
        for position, statement in enumerate(items):
            statements.append(
                {
                    "stmt_id": f"{sao_id}#{position}",
                    "sao_id": sao_id,
                    "position": position,
                    "filing_year": int(row["filing_year"]),
                    "_recovered": bool(recovered.loc[sao_id]),
                    "verbatim": verbatims[position],
                    "prev_verbatim": (
                        clip(verbatims[position - 1], CONTEXT_CHARS)
                        if position > 0
                        else ""
                    ),
                    "next_verbatim": (
                        clip(verbatims[position + 1], CONTEXT_CHARS)
                        if position + 1 < len(verbatims)
                        else ""
                    ),
                    **{field: statement.get(field) for field in STRATUM_FIELDS},
                }
            )
        context, context_source = rmad_context(row[text_col])
        documents.append(
            {
                "filing_id": sao_id,
                "filing_year": int(row["filing_year"]),
                "_recovered": bool(recovered.loc[sao_id]),
                "rmad_conclusion": record.get("rmad_conclusion"),
                "rmad_sentence_context": context,
                "context_source": context_source,
            }
        )
    logger.info(
        "Loaded %s statements across %s filings.",
        f"{len(statements):,}",
        f"{len(documents):,}",
    )
    return pd.DataFrame(statements), pd.DataFrame(documents)


def window(text: str, start: int, length: int) -> str:
    """Take a context window of roughly `RMAD_CONTEXT_CHARS` around a span.

    Args:
        text: The display text.
        start: Index the span begins at.
        length: Length of the span.

    Returns:
        A new substring, marked with an ellipsis where it was cut.
    """
    half = max(0, (RMAD_CONTEXT_CHARS - length) // 2)
    lo = max(0, start - half)
    hi = min(len(text), start + length + half)
    return ("… " if lo > 0 else "") + text[lo:hi] + (" …" if hi < len(text) else "")


def rmad_context(narrative: str) -> tuple[str, str]:
    """Return narrative text surrounding the RMAD discussion.

    ⚠️ This deliberately does **not** consult the model's `rmad_verbatim`. Centring the
    window on the extracted sentence makes the window's very shape a function of the
    model's output: every filing the model called `not_stated` has no sentence to
    centre on, so it would receive a visibly different kind of context, and a coder who
    noticed the pattern would be nudged toward `not_stated`. Building every window the
    same way removes that channel. On the drawn sample this locates the passage for 136
    of 150 documents without reference to any model output.

    The remaining 14 fall back to the head of the narrative because the filing never
    mentions the subject at all. That is a property of the source text rather than a
    hint from the model: a coder seeing no discussion of material adverse deviation
    should conclude `not_stated`, and would do so from the document itself.

    Locating the passage needs the same normalisation the quotation check uses -- page
    furniture is injected mid-sentence and hyphenated compounds split across lines, so a
    naive search misses passages that are plainly present. The displayed text is the
    furniture-stripped narrative with whitespace collapsed rather than the normalised
    one, so the actuary's own characters survive; the match index is scaled between the
    two, drifting by a few characters at most inside a 600-character window.

    Args:
        narrative: The full narrative.

    Returns:
        A tuple of (context, source) where source is "keyword" or "head".
    """
    stripped = strip_furniture(str(narrative))
    display = squash(stripped)
    search = normalize_punct(stripped)
    if not search:
        return "", "head"
    scale = len(display) / len(search)

    lowered = search.lower()
    for keyword in RMAD_KEYWORDS:
        index = lowered.find(keyword)
        if index >= 0:
            start = min(len(display) - 1, int(index * scale))
            return window(display, start, int(len(keyword) * scale)), "keyword"
    return display[:RMAD_CONTEXT_CHARS] + (" …" if len(display) > RMAD_CONTEXT_CHARS else ""), "head"


def assign_strata(statements: pd.DataFrame) -> pd.Series:
    """Label each statement with its sampling stratum.

    Cells below `RARE_THRESHOLD` of the population are pooled into one `RARE`
    stratum. Dropping them instead would give 1,156 statements a zero inclusion
    probability and quietly restrict the estimand; sampling each separately would
    triple the weight spread for cells too small to estimate anything.

    Args:
        statements: The statement frame.

    Returns:
        A new Series of stratum labels.
    """
    key = statements[STRATUM_FIELDS].agg(" | ".join, axis=1)
    sizes = key.value_counts()
    rare = set(sizes.index[sizes < RARE_THRESHOLD * len(statements)])
    logger.info(
        "Strata: %d cells, %d pooled into RARE (%s statements).",
        len(sizes),
        len(rare),
        f"{int(sizes[list(rare)].sum()):,}",
    )
    return key.where(~key.isin(rare), "RARE")


def largest_remainder(counts: pd.Series, total: int) -> pd.Series:
    """Apportion `total` across groups proportionally, without rounding drift.

    Args:
        counts: Population size per group.
        total: How many to allocate in all.

    Returns:
        A new integer Series summing to `min(total, counts.sum())`.
    """
    total = int(min(total, counts.sum()))
    exact = counts / counts.sum() * total
    base = np.floor(exact).astype(int)
    base = np.minimum(base, counts)
    remainder = total - int(base.sum())
    if remainder > 0:
        order = (exact - base).sort_values(ascending=False).index
        for group in order:
            if remainder == 0:
                break
            if base[group] < counts[group]:
                base[group] += 1
                remainder -= 1
    return base


def draw(frame: pd.DataFrame, allocation: pd.Series, seed: int) -> pd.DataFrame:
    """Sample within each stratum, balancing across filing years.

    Inclusion probability is computed at the (stratum, year) level rather than the
    stratum level: year balancing makes it vary within a stratum, and treating it as
    constant would bias every corrected estimate.

    Args:
        frame: The population frame, carrying `stratum` and `filing_year`.
        allocation: Rows to draw per stratum.
        seed: Random seed.

    Returns:
        A new DataFrame of sampled rows with `pi` and `weight`.
    """
    drawn = []
    for stratum, take in allocation.items():
        pool = frame.loc[frame["stratum"] == stratum]
        per_year = largest_remainder(pool["filing_year"].value_counts(), int(take))
        for year, count in per_year.items():
            if count == 0:
                continue
            cell = pool.loc[pool["filing_year"] == year]
            picked = cell.sample(n=int(count), random_state=seed)
            pi = int(count) / len(cell)
            drawn.append(
                picked.assign(
                    stratum=stratum,
                    stratum_year=f"{stratum} @ {year}",
                    pi=pi,
                    weight=1.0 / pi,
                )
            )
    return pd.concat(drawn, ignore_index=True)


def allocate_statements(statements: pd.DataFrame) -> pd.Series:
    """Allocate the statement sample across strata.

    Args:
        statements: The statement frame carrying `stratum`.

    Returns:
        A new Series of rows to draw per stratum.
    """
    sizes = statements["stratum"].value_counts()
    return pd.Series(
        np.minimum(np.minimum(CELL_FLOOR, sizes.values), CELL_CAP), index=sizes.index
    )


def allocate_documents(documents: pd.DataFrame) -> pd.Series:
    """Allocate the document sample roughly evenly across RMAD classes.

    Args:
        documents: The document frame.

    Returns:
        A new Series of rows to draw per class.
    """
    sizes = documents["stratum"].value_counts()
    # Even across classes, not proportional: `no` outnumbers `conditional` 67:1, and a
    # proportional draw would leave the rare conclusions unestimable. largest_remainder
    # is not usable here -- it apportions a total across populations, whereas this
    # splits the total evenly and only then caps by population.
    count = len(sizes)
    base = np.full(count, DOC_TARGET // count, dtype=int)
    base[: DOC_TARGET % count] += 1
    return pd.Series(np.minimum(base, sizes.values), index=sizes.index)


def estimate_height(text: str, width: int, cap: int = 240) -> float:
    """Approximate the row height needed to show wrapped text.

    openpyxl cannot autofit, so height is estimated from length. Slightly generous,
    because a clipped verbatim is worse than a tall row.

    Args:
        text: The cell text.
        width: Column width in characters.
        cap: Maximum height in points.

    Returns:
        A row height in points.
    """
    lines = max(1, int(np.ceil(len(text) / max(width - 2, 10))))
    return float(min(cap, 14 + 12.5 * lines))


def write_sheet(
    workbook: Workbook,
    title: str,
    columns: list[str],
    coding: list[str],
    rows: list[dict[str, Any]],
    widths: dict[str, int],
    wrap_column: str,
    freeze: str,
    example: dict[str, Any],
) -> None:
    """Write one coding sheet, formatted, validated and protected.

    Args:
        workbook: The workbook to add the sheet to.
        title: Sheet title.
        columns: Column order.
        coding: Columns the coder may edit.
        rows: Data rows.
        widths: Column width per column name.
        wrap_column: The column whose height drives row height.
        freeze: Freeze-pane anchor.
        example: The worked EXAMPLE row.
    """
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for index, name in enumerate(columns, 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths[name]
        header = sheet.cell(row=1, column=index)
        header.font = Font(bold=True)
        header.fill = HEADER_FILL
        header.alignment = Alignment(vertical="center", wrap_text=True)

    for record in [example, *rows]:
        sheet.append([xl_safe(record.get(name, "")) for name in columns])

    for row_index in range(2, sheet.max_row + 1):
        text = str(sheet.cell(row=row_index, column=columns.index(wrap_column) + 1).value or "")
        sheet.row_dimensions[row_index].height = estimate_height(
            text, widths[wrap_column]
        )
        for column_index, name in enumerate(columns, 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(
                vertical="top", wrap_text=name in {wrap_column, "prev_verbatim",
                                                   "next_verbatim", "notes",
                                                   "rmad_sentence_context"}
            )
            if name in coding:
                cell.fill = CODING_FILL
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)
        if row_index == 2:
            for column_index in range(1, len(columns) + 1):
                sheet.cell(row=row_index, column=column_index).fill = EXAMPLE_FILL

    for name, options in CODES.items():
        if name not in columns:
            continue
        letter = get_column_letter(columns.index(name) + 1)
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        validation.error = "Choose a value from the list."
        validation.prompt = "Select one: " + ", ".join(options)
        sheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{sheet.max_row}")

    sheet.freeze_panes = freeze
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"


def write_codebook_sheet(workbook: Workbook, path: Path) -> str:
    """Paste the codebook verbatim into its own sheet, stamped with its hash.

    A workbook has to say which instrument it carries: the coder follows Sheet 3, and
    a workbook built against a superseded codebook is worse than no workbook.

    Args:
        workbook: The workbook to add the sheet to.
        path: Path to the codebook markdown.

    Returns:
        The codebook's short hash.
    """
    text = path.read_text(encoding="utf-8")
    digest = hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]
    sheet = workbook.create_sheet("Codebook")
    sheet.column_dimensions["A"].width = 110
    stamp = f"Instrument: {path.name} @ {digest}"
    for line in [stamp, ""] + text.splitlines():
        sheet.append([xl_safe(line)])
    sheet.cell(row=1, column=1).font = Font(bold=True)
    for row_index in range(1, sheet.max_row + 1):
        sheet.cell(row=row_index, column=1).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
    sheet.protection.sheet = True
    return digest


def write_design_sheet(
    workbook: Workbook, statements: pd.DataFrame, documents: pd.DataFrame,
    very_hidden: bool,
) -> None:
    """Write the hidden design sheet holding ids, strata and weights.

    ⚠️ This sheet contains the stratum, which is the model's label triple. The
    workbook is blind only while the sheet stays hidden.

    Args:
        workbook: The workbook to add the sheet to.
        statements: The drawn statement rows.
        documents: The drawn document rows.
        very_hidden: Whether to mark the sheet veryHidden rather than hidden.
    """
    sheet = workbook.create_sheet("Design")
    columns = ["level", "id", "stratum", "stratum_year", "pi", "weight", "_recovered"]
    sheet.append(columns)
    for _, row in statements.iterrows():
        sheet.append(
            ["statement", row["stmt_id"], row["stratum"], row["stratum_year"],
             row["pi"], row["weight"], row["_recovered"]]
        )
    for _, row in documents.iterrows():
        sheet.append(
            ["document", row["filing_id"], row["stratum"], row["stratum_year"],
             row["pi"], row["weight"], row["_recovered"]]
        )
    sheet.sheet_state = "veryHidden" if very_hidden else "hidden"


def build_example(kind: str) -> dict[str, Any]:
    """Build the worked EXAMPLE row for a sheet.

    The coding cells are left blank on purpose. Filling them would mean scoring a
    statement, which is the annotator's job, and a machine-suggested answer in a
    human column is exactly what this workbook exists to avoid. The row demonstrates
    the layout and the dropdowns; the codebook supplies the judgement.

    Args:
        kind: Either "statement" or "document".

    Returns:
        A new dict for the example row.
    """
    if kind == "statement":
        return {
            "stmt_id": "EXAMPLE — not part of the draw",
            "filing_year": "",
            "prev_verbatim": "(the statement immediately before this one, for the "
                             "codebook's adjacency rule)",
            "verbatim": "(the full quoted statement appears here, wrapped so it can "
                        "be read without clicking into the cell)",
            "next_verbatim": "(the statement immediately after this one)",
            "orientation": "",
            "object": "",
            "moment": "",
            "notes": "Type in the shaded cells; each has a dropdown. Left blank on "
                     "purpose — no suggested coding.",
        }
    return {
        "filing_id": "EXAMPLE — not part of the draw",
        "filing_year": "",
        "rmad_sentence_context": "(~600 characters of the narrative around the RMAD "
                                 "conclusion, so the actuary's own sentence is read "
                                 "in context)",
        "rmad_conclusion": "",
        "notes": "Type in the shaded cells; the dropdown lists the four values.",
    }


def main() -> None:
    """Draw both samples and write the workbook and draw index."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dir", default="raw_full", help="extraction directory")
    parser.add_argument(
        "--codebook",
        default="instrument/codebook_v1.0.md",
        help="codebook markdown pasted into Sheet 3",
    )
    parser.add_argument(
        "--out", default="validation_blind_v1.0.xlsx", help="workbook filename"
    )
    parser.add_argument(
        "--index", default="validation_draw.csv", help="draw index filename"
    )
    parser.add_argument(
        "--very-hidden",
        action="store_true",
        help="mark the design sheet veryHidden so it cannot be unhidden from the UI",
    )
    parser.add_argument("--force", action="store_true", help="overwrite an existing workbook")
    args = parser.parse_args()

    codebook = Path(args.codebook)
    if not codebook.is_absolute():
        codebook = BASE_DIR / codebook
    if not codebook.exists():
        raise SystemExit(
            f"Codebook not found at {codebook}. Sheet 3 must carry the instrument the "
            "coder follows, and falling back to another file would have them coding "
            "to superseded rules. Write it, or pass --codebook explicitly."
        )

    target = BASE_DIR / args.out
    if target.exists() and not args.force:
        raise SystemExit(f"{target} already exists; pass --force to overwrite.")

    out_dir = DEFAULT_OUT_ROOT / args.dir
    statements, documents = load_frames(out_dir)

    statements["stratum"] = assign_strata(statements)
    documents["stratum"] = documents["rmad_conclusion"].astype(str)

    stmt_draw = draw(statements, allocate_statements(statements), RANDOM_STATE)
    doc_draw = draw(documents, allocate_documents(documents), RANDOM_STATE)

    stmt_draw = stmt_draw.sort_values(["sao_id", "position"], ignore_index=True)
    doc_draw = doc_draw.sort_values(["filing_id"], ignore_index=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(
        workbook, "Statements", STMT_COLUMNS, STMT_CODING,
        stmt_draw[STMT_COLUMNS[:5]].to_dict("records"),
        {"stmt_id": 26, "filing_year": 10, "prev_verbatim": 40, "verbatim": 80,
         "next_verbatim": 40, "orientation": 16, "object": 18, "moment": 14,
         "notes": 34},
        "verbatim", "F2", build_example("statement"),
    )
    write_sheet(
        workbook, "Documents", DOC_COLUMNS, DOC_CODING,
        doc_draw[DOC_COLUMNS[:3]].to_dict("records"),
        {"filing_id": 30, "filing_year": 10, "rmad_sentence_context": 100,
         "rmad_conclusion": 18, "notes": 34},
        "rmad_sentence_context", "D2", build_example("document"),
    )
    digest = write_codebook_sheet(workbook, codebook)
    write_design_sheet(workbook, stmt_draw, doc_draw, args.very_hidden)
    workbook.save(target)

    index = pd.concat(
        [
            stmt_draw.assign(level="statement").rename(columns={"stmt_id": "id"})[
                ["level", "id", "sao_id", "filing_year", "stratum", "stratum_year",
                 "pi", "weight", "_recovered"]
            ],
            doc_draw.assign(level="document", sao_id=doc_draw["filing_id"]).rename(
                columns={"filing_id": "id"}
            )[
                ["level", "id", "sao_id", "filing_year", "stratum", "stratum_year",
                 "pi", "weight", "_recovered", "context_source"]
            ],
        ],
        ignore_index=True,
    )
    index.to_csv(BASE_DIR / args.index, index=False)

    logger.info("Wrote %s (codebook %s) and %s.", target.name, digest, args.index)
    report(stmt_draw, doc_draw, statements, documents)


def report(
    stmt_draw: pd.DataFrame, doc_draw: pd.DataFrame,
    statements: pd.DataFrame, documents: pd.DataFrame,
) -> None:
    """Print realised stratum counts and weight ranges for a design check.

    Args:
        stmt_draw: The drawn statement rows.
        doc_draw: The drawn document rows.
        statements: The statement population.
        documents: The document population.
    """
    for name, drawn, population in [
        ("STATEMENTS", stmt_draw, statements),
        ("DOCUMENTS", doc_draw, documents),
    ]:
        sizes = population["stratum"].value_counts()
        table = (
            drawn.groupby("stratum")
            .agg(drawn=("pi", "size"), weight_min=("weight", "min"),
                 weight_max=("weight", "max"))
            .join(sizes.rename("population"))
        )
        table = table[["population", "drawn", "weight_min", "weight_max"]]
        logger.info("\n--- %s: n=%d across %d strata ---\n%s",
                    name, len(drawn), len(table), table.round(1).to_string())
        # Horvitz-Thompson: the weights must reconstruct the population total. Any
        # drift here means pi was computed at the wrong level, which would bias every
        # corrected estimate downstream.
        covered = set(zip(drawn["stratum"], drawn["filing_year"]))
        uncovered = int(
            (~pd.Series(
                list(zip(population["stratum"], population["filing_year"]))
            ).isin(covered)).sum()
        )
        logger.info(
            "%s weights %.1f-%.1f | sum(weight)=%.0f vs population %d | "
            "rows with zero inclusion probability: %d",
            name, drawn["weight"].min(), drawn["weight"].max(),
            drawn["weight"].sum(), len(population), uncovered,
        )


if __name__ == "__main__":
    main()
